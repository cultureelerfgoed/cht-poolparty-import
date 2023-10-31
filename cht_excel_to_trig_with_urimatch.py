from openpyxl import load_workbook
from rdflib import Graph, Namespace, Literal, URIRef, plugin
from rdflib.namespace import RDF, SKOS
import uuid
from rdflib import query
import requests


# Create an RDF graph
g = Graph()

# Define namespaces
CULTUREELERFGOED = Namespace("https://data.cultureelerfgoed.nl/term/id/cht/")
SEMANTIC_WEB = Namespace("http://schema.semantic-web.at/ppt/")

# Define the API endpoint
API_ENDPOINT = "https://api.linkeddata.cultureelerfgoed.nl/queries/rce/cht-uri-fetch-pp/run?"

# Function to generate the RDF triples for a row
def create_rdf_triples(row):
  #  broader_term_uri = URIRef(f"{row['broader_uri']}")
    generated_uuid = str(uuid.uuid4())  # Generate a random UUID
    term_uri = URIRef(f"{CULTUREELERFGOED}{generated_uuid}")

# Initialize related_uri with None
    related_uri = None

# Lookup URI for the "broader_term"
    broader_uri = fetch_uri(row['broader_term'])

    # Create the RDF triples in the specified graph
    g.add((term_uri, RDF.type, SKOS.Concept))
    g.add((term_uri, SEMANTIC_WEB.inSubtree, URIRef(broader_uri)))
    g.add((term_uri, SKOS.inScheme, CULTUREELERFGOED["b532325c-dc08-49db-b4f1-15e53b037ec3"]))
    g.add((CULTUREELERFGOED["b532325c-dc08-49db-b4f1-15e53b037ec3"], RDF.type, SKOS.ConceptScheme))
    g.add((term_uri, SKOS.broader, URIRef(broader_uri)))
    g.add((term_uri, SKOS.prefLabel, Literal(row['term'], lang="nl")))
    g.add((term_uri, SKOS.scopeNote, Literal(row['beschrijving'], lang="nl")))
    
    # Handle multiple altLabels
    if "|" in row['alternatief']:
        alt_labels = row['alternatief'].split("|")
        for alt_label in alt_labels:
            alt_label = alt_label.strip()  # Remove leading/trailing spaces
            g.add((term_uri, SKOS.altLabel, Literal(alt_label, lang="nl")))
    else:
        g.add((term_uri, SKOS.altLabel, Literal(row['alternatief'], lang="nl")))
    # Handle multiple related terms
    related_terms = row['gerelateerd'].split("|")
    for related_term in related_terms:
        related_uri = fetch_uri(related_term)
        if related_uri:
            g.add((term_uri, SKOS.related, URIRef(related_uri)))

# Split the "gerelateerd" string by "|"
    related_terms = row['gerelateerd'].split("|")

    # Query the API to fetch the URI for each related term
    for related_term in related_terms:
        related_uri = fetch_uri(related_term)
        if related_uri:
            g.add((term_uri, SKOS.related, URIRef(related_uri)))

# Function to make an API call to fetch the URI for a term
def fetch_uri(term):
    endpoint = API_ENDPOINT  # Use the same endpoint for both broader and related terms
    params = {"term": term}
    response = requests.get(endpoint, params=params)
    if response.status_code == 200:
        data = response.json()
        if isinstance(data, list) and len(data) > 0:
            first_item = data[0]
            return first_item.get("uri")
    return None

# Load the XLSX file
xlsx_file = 'C:\\Users\\Ruben\\Documents\\05. RCE\\CHT\\cht-poolparty-import\\test_import_pp.xlsx'
wb = load_workbook(filename=xlsx_file, read_only=True)
sheet = wb.active

# Iterate through rows in the worksheet and create RDF triples
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_data = {
        'broader_term': row[1],  # Adjust column index as needed
        'term': row[3],         # Adjust column index as needed
        'beschrijving': row[4],  # Adjust column index as needed
        'alternatief': row[5],   # Adjust column index as needed
        'gerelateerd': row[6]    # Adjust column index as needed
    }
    create_rdf_triples(row_data)

# Serialize the RDF graph to Turtle format
ttl_output = g.serialize(format='turtle')

# Save the TTL data to a file
with open('output.ttl', 'w') as output_file:
    output_file.write(ttl_output)

print("TTL conversion completed. Check 'output.ttl'.")
