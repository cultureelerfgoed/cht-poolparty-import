from openpyxl import load_workbook
from rdflib import Graph, Namespace, Literal, URIRef
from rdflib.namespace import RDF, SKOS
import uuid  

# Create an RDF graph
g = Graph()

# Define namespaces
CULTUREELERFGOED = Namespace("https://data.cultureelerfgoed.nl/term/id/cht/")
SEMANTIC_WEB = Namespace("http://schema.semantic-web.at/ppt/")

# Function to generate the RDF triples for a row
def create_rdf_triples(row):
    broader_term_uri = URIRef(f"{row['broader_uri']}")
    generated_uuid = str(uuid.uuid4())  # Generate a random UUID
    term_uri = URIRef(f"{CULTUREELERFGOED}{generated_uuid}")

    # Create the RDF triples
    g.add((term_uri, RDF.type, SKOS.Concept))
    g.add((term_uri, SEMANTIC_WEB.inSubtree, broader_term_uri))
    g.add((term_uri, SKOS.inScheme, CULTUREELERFGOED["b532325c-dc08-49db-b4f1-15e53b037ec3"]))
    g.add((CULTUREELERFGOED["b532325c-dc08-49db-b4f1-15e53b037ec3"], RDF.type, SKOS.ConceptScheme))
    g.add((term_uri, SKOS.broader, broader_term_uri))
    g.add((broader_term_uri, SKOS.narrower, term_uri))
    g.add((term_uri, SKOS.prefLabel, Literal(row['term'], lang="nl")))
    g.add((term_uri, SKOS.scopeNote, Literal(row['beschrijving'], lang="nl")))
        # add related terms with URI's if possible; add concept status

        
         # Handle multiple altLabels
    if "|" in row['alternatief']:
        alt_labels = row['alternatief'].split("|")
        for alt_label in alt_labels:
            alt_label = alt_label.strip()  # Remove leading/trailing spaces
            g.add((term_uri, SKOS.altLabel, Literal(alt_label, lang="nl")))
    else:
        g.add((term_uri, SKOS.altLabel, Literal(row['alternatief'], lang="nl")))

# Load the XLSX file
xlsx_file = 'C:\\Users\\Ruben\\Documents\\05. RCE\\CHT\\cht-poolparty-import\\test_import_pp.xlsx'
wb = load_workbook(filename=xlsx_file, read_only=True)
sheet = wb.active

# Iterate through rows in the worksheet and create RDF triples
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_data = {
        'broader_uri': row[2],  # Adjust column index as needed
        'term': row[3],         # Adjust column index as needed
        'beschrijving': row[4],  # Adjust column index as needed
        'alternatief': row[5]    # Adjust column index as needed
    }
    create_rdf_triples(row_data)

# Serialize the RDF graph to Turtle format
ttl_output = g.serialize(format='turtle')

# Save the TTL data to a file
with open('output.ttl', 'w') as output_file:
    output_file.write(ttl_output)

print("TTL conversion completed. Check 'output.ttl'.")
